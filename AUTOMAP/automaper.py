"""
automaper.py

Utilities for loading, visualizing and managing
photonic chip layouts.
"""

import pandas as pd
from pprint import pprint
from openpyxl import load_workbook

# ======================================================
# EXCEPTIONS
# ======================================================

class MeasurementPlanGenerated(Exception):
    """
    Raised when AUTOMAPER generates a new measurement plan.

    The user must review the Measurements sheet before
    running the program again.
    """
    pass


# ======================================================
# LOADERS
# ======================================================

def load_chip(excel_file):
    """
    Load a photonic chip from an Excel file.

    If the workbook does not contain a Measurements sheet,
    AUTOMAPER automatically generates one, saves the Excel
    file and stops the execution so the user can review it.
    """

    # --------------------------------------------------
    # Read mandatory sheets
    # --------------------------------------------------

    metadata_df = pd.read_excel(
        excel_file,
        sheet_name="Metadata"
    )

    devices_df = pd.read_excel(
        excel_file,
        sheet_name="Devices"
    )

    ports_df = pd.read_excel(
        excel_file,
        sheet_name="Ports"
    )

    # --------------------------------------------------
    # Build chip object
    # --------------------------------------------------

    chip = _build_chip(
        devices_df,
        ports_df
    )

    # --------------------------------------------------
    # Check if Measurements sheet exists
    # --------------------------------------------------

    workbook = load_workbook(excel_file)

    if "MeasurementPaths" not in workbook.sheetnames:

        _generate_measurement_paths_sheet(
            excel_file,
            chip
        )

        raise MeasurementPlanGenerated(

            "\n"
            "No measurement paths were found.\n\n"

            "AUTOMAPER has automatically generated a "
            "'MeasurementPaths' sheet.\n\n"

            "Please:\n"
            "  1. Open the Excel file.\n"
            "  2. Review the generated measurement paths.\n"
            "  3. Enable/disable the desired measurements.\n"
            "  4. Save the Excel file.\n"
            "  5. Run the program again.\n"

        )

    # --------------------------------------------------
    # Read measurement plan
    # --------------------------------------------------

    measurement_paths_df = pd.read_excel(
        excel_file,
        sheet_name="MeasurementPaths"
    )

    measurement_paths = _build_measurement_paths(
        measurement_paths_df
    )

    measurement_plan = create_measurement_plan(
    chip,
    measurement_paths
    )

    return chip, measurement_paths, measurement_plan

# ======================================================
# VISUALIZATION
# ======================================================

def plot_chip(chip):
    """
    Display the loaded chip.
    """

    print("=" * 80)
    print("CHIP")
    print("=" * 80)
    pprint(chip)


def plot_measurement_paths(measurement_paths, chip):
    """
    Display the enabled measurement plan.
    """

    print("=" * 80)
    print("MEASUREMENT PATHS")
    print("=" * 80)

    for measurement in measurement_paths:

        device = chip[measurement["device"]]

        input_port = device["ports"][measurement["input"]]

        output_port = device["ports"][measurement["output"]]

        print("-" * 80)

        print(f"Device : {measurement['device']}")

        print(
            f"Input  : {measurement['input']}"
            f" ({input_port['x']:.2f}, {input_port['y']:.2f}) µm"
        )

        print(
            f"Output : {measurement['output']}"
            f" ({output_port['x']:.2f}, {output_port['y']:.2f}) µm"
        )

        # print(f"Status : {measurement['status']}")


def plot_measurement_plan(measurement_plan):
    """
    Display the generated measurement plan.
    """

    print("=" * 80)
    print("MEASUREMENT PLAN")
    print("=" * 80)

    for measurement in measurement_plan:

        print("-" * 80)

        print(f"Device : {measurement['device']}")

        print(
            f"Input  : {measurement['input']['name']}"
            f" ({measurement['input']['x']:.2f}, "
            f"{measurement['input']['y']:.2f}) µm"
        )

        print(
            f"Output : {measurement['output']['name']}"
            f" ({measurement['output']['x']:.2f}, "
            f"{measurement['output']['y']:.2f}) µm"
        )

        print(f"Status : {measurement['status']}")


# def save_measurement_results(chip, filename):
#     """
#     Guarda los resultados de las medidas.
#     """

#     pass


# def export_csv(chip, filename):
#     """
#     Exporta los resultados a CSV.
#     """

#     pass

# ======================================================
# FUNCIONES
# ======================================================

def create_measurement_plan(chip, measurement_paths):
    """
    Create the measurement plan from the chip description
    and the enabled measurement paths.
    """

    measurement_plan = []

    for path in measurement_paths:

        # ----------------------------------------------
        # Get device
        # ----------------------------------------------

        device = chip[path["device"]]

        # ----------------------------------------------
        # Skip disabled devices
        # ----------------------------------------------

        if not device["measure"]:
            continue

        # ----------------------------------------------
        # Skip disabled paths
        # ----------------------------------------------

        if not path["enabled"]:
            continue

        # ----------------------------------------------
        # Get ports
        # ----------------------------------------------

        input_port = device["ports"][path["input"]]

        output_port = device["ports"][path["output"]]

        # ----------------------------------------------
        # Create measurement
        # ----------------------------------------------

        measurement_plan.append({

            "device": path["device"],

            "input": {

                "name": path["input"],

                "x": input_port["x"],

                "y": input_port["y"]

            },

            "output": {

                "name": path["output"],

                "x": output_port["x"],

                "y": output_port["y"]

            },

            "status": path["status"],

            "power": path["power"],

            "timestamp": path["timestamp"]

        })

    return measurement_plan

# ======================================================
# FUNCIONES INTERNAS
# ======================================================

def _build_chip(devices_df, ports_df):

    chip = {}

    # --------------------------------------------------
    # Add devices
    # --------------------------------------------------

    for _, row in devices_df.iterrows():

        chip[row["Device"]] = {

            "type": row["Type"],

            "measure": row["Measure"] == "YES",

            "ports": {}

        }

    # --------------------------------------------------
    # Add ports
    # --------------------------------------------------

    for _, row in ports_df.iterrows():

        device = row["Device"]

        chip[device]["ports"][row["Port"]] = {

            "x": row["X (um)"],

            "y": row["Y (um)"],

            "direction": row["Direction"]

        }

    return chip


def _build_measurement_paths(measurement_paths_df):

    measurement_paths = []

    for _, row in measurement_paths_df.iterrows():

        measurement_paths.append({

            "device": row["Device"],

            "input": row["Input"],

            "output": row["Output"],

            "enabled": row["Enabled"] == "YES",

            "status": row["Status"],

            "power": None,

            "timestamp": None

        })

    return measurement_paths


def _generate_measurement_paths_sheet(excel_file, chip):
    """
    Automatically generate the MeasurementPaths sheet.

    All possible Input -> Output combinations are created
    and enabled by default.
    """

    workbook = load_workbook(excel_file)

    if "MeasurementPaths" in workbook.sheetnames:
        del workbook["MeasurementPaths"]

    worksheet = workbook.create_sheet("MeasurementPaths")

    worksheet.append([

        "Device",

        "Input",

        "Output",

        "Enabled",

        "Status"

    ])

    # --------------------------------------------------
    # Generate all Input -> Output combinations
    # --------------------------------------------------

    for device_name, device in chip.items():

        inputs = []
        outputs = []

        for port_name, port in device["ports"].items():

            if port["direction"] == "IN":

                inputs.append(port_name)

            elif port["direction"] == "OUT":

                outputs.append(port_name)

        for input_port in inputs:

            for output_port in outputs:

                worksheet.append([

                    device_name,

                    input_port,

                    output_port,

                    "YES",

                    "Pending"

                ])

    workbook.save(excel_file)