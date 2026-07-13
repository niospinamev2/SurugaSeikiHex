"""
automaper.py

Utilities for loading, visualizing and managing
photonic chip layouts.
"""

import pandas as pd
from pprint import pprint
from openpyxl import load_workbook

# ======================================================
# EXCEPTIONS
# ======================================================

class MeasurementPlanGenerated(Exception):
    """
    Raised when AUTOMAPER generates a new measurement plan.

    The user must review the Measurements sheet before
    running the program again.
    """
    pass


# ======================================================
# LOADERS
# ======================================================

def load_chip(excel_file):
    """
    Load a photonic chip from an Excel file.

    If the workbook does not contain a Measurements sheet,
    AUTOMAPER automatically generates one, saves the Excel
    file and stops the execution so the user can review it.
    """

    # --------------------------------------------------
    # Read mandatory sheets
    # --------------------------------------------------

    metadata_df = pd.read_excel(
        excel_file,
        sheet_name="Metadata"
    )

    devices_df = pd.read_excel(
        excel_file,
        sheet_name="Devices"
    )

    ports_df = pd.read_excel(
        excel_file,
        sheet_name="Ports"
    )

    # --------------------------------------------------
    # Build chip object
    # --------------------------------------------------

    chip = _build_chip(
        devices_df,
        ports_df
    )

    # --------------------------------------------------
    # Check if Measurements sheet exists
    # --------------------------------------------------

    workbook = load_workbook(excel_file)

    if "Measurements" not in workbook.sheetnames:

        _generate_measurement_plan_sheet(
            excel_file,
            chip
        )

        raise MeasurementPlanGenerated(

            "\n"
            "No measurement plan was found.\n\n"

            "AUTOMAPER has automatically generated a "
            "'Measurements' sheet.\n\n"

            "Please:\n"
            "  1. Open the Excel file.\n"
            "  2. Review the generated measurement plan.\n"
            "  3. Enable/disable the desired measurements.\n"
            "  4. Save the Excel file.\n"
            "  5. Run the program again.\n"

        )

    # --------------------------------------------------
    # Read measurement plan
    # --------------------------------------------------

    measurements_df = pd.read_excel(
        excel_file,
        sheet_name="Measurements"
    )

    measurement_plan = _build_measurement_plan(
        measurements_df
    )

    return chip, measurement_plan

# ======================================================
# VISUALIZATION
# ======================================================

def plot_chip(chip):
    """
    Display the loaded chip.
    """

    print("=" * 80)
    print("CHIP")
    print("=" * 80)
    pprint(chip)


# def plot_measurement_plan(measurement_plan):
#     """
#     Muestra el plan de medida.
#     """

#     pass


# def save_measurement_results(chip, filename):
#     """
#     Guarda los resultados de las medidas.
#     """

#     pass


# def export_csv(chip, filename):
#     """
#     Exporta los resultados a CSV.
#     """

#     pass


# ======================================================
# FUNCIONES INTERNAS
# ======================================================

def _build_chip(devices_df, ports_df):

    chip = {}

    # --------------------------------------------------
    # Add devices
    # --------------------------------------------------

    for _, row in devices_df.iterrows():

        chip[row["Device"]] = {

            "type": row["Type"],

            "measure": row["Measure"],

            "ports": {}

        }

    # --------------------------------------------------
    # Add ports
    # --------------------------------------------------

    for _, row in ports_df.iterrows():

        device = row["Device"]

        chip[device]["ports"][row["Port"]] = {

            "x": row["X (um)"],

            "y": row["Y (um)"],

            "direction": row["Direction"]

        }

    return chip


def _build_measurement_plan(measurements_df):

    measurement_plan = []

    for _, row in measurements_df.iterrows():

        measurement_plan.append({

            "device": row["Device"],

            "input": row["Input"],

            "output": row["Output"],

            "enabled": row["Enabled"] == "YES",

            "status": row["Status"],

            "power": None,

            "timestamp": None

        })

    return measurement_plan


def _generate_measurement_plan_sheet(excel_file, chip):
    """
    Automatically generate the Measurements sheet.

    All possible Input -> Output combinations are created
    and enabled by default.
    """

    workbook = load_workbook(excel_file)

    if "Measurements" in workbook.sheetnames:
        del workbook["Measurements"]

    worksheet = workbook.create_sheet("Measurements")

    worksheet.append([

        "Device",

        "Input",

        "Output",

        "Enabled",

        "Status"

    ])

    # --------------------------------------------------
    # Generate all Input -> Output combinations
    # --------------------------------------------------

    for device_name, device in chip.items():

        inputs = []
        outputs = []

        for port_name, port in device["ports"].items():

            if port["direction"] == "IN":

                inputs.append(port_name)

            elif port["direction"] == "OUT":

                outputs.append(port_name)

        for input_port in inputs:

            for output_port in outputs:

                worksheet.append([

                    device_name,

                    input_port,

                    output_port,

                    "YES",

                    "Pending"

                ])

    workbook.save(excel_file)